import openpyxl
import csv
import sys
import os.path
import io

XPATH = './data.xlsx'
CPATH = './data.csv'
OUT = 'out.xlsx'

def fexist(path = '.') :
    check_file = os.path.isfile(path)
    return check_file

class data:
    __wb = None
    __ws = None
    __columns = {}
    def __init__(self):

        if fexist(XPATH): #if data.xlsx in folder -> load
            self.__wb = openpyxl.load_workbook(XPATH)
            self.__ws = self.__wb.active
            print("excel data loaded...")
            print(f"Total number of rows: {self.__ws.max_row}. And total number of columns: {self.__ws.max_column}!")
            self.__post_load()

        elif fexist(CPATH): #if data.csv in folder -> convert, load
            self.__wb = openpyxl.Workbook()
            self.__ws = self.__wb.active
            with io.open(CPATH, encoding='utf-8') as f:
                reader = csv.reader(f, delimiter=',')
                for row in reader:
                    self.__ws.append(row)
            print("csv data converted to excel and loaded...")
            print(f"Total number of rows: {self.__ws.max_row}. And total number of columns: {self.__ws.max_column}!")
            self.__post_load()

        else: #if no data.* in folder -> exit
            sys.exit("no data found!")



    def __post_load(self): #search for some data to remember for faster work after excel is loaded
        self.__load_columns()
        self.set_column_size(30)

    def __load_columns(self): #remember columns id and names
        for i in range(1, self.__ws.max_column + 1):
            col_name = self.__ws[str(chr(64 + i))+"1"].value
            if col_name == None:
                break
            self.__columns[col_name] = i

    def __apply_col_del(self): #after deliting an element u need to apply ids to all keys again
        j = 1
        for i in self.__columns:
            self.__columns[i] = j
            j += 1

    def __get_list_from_row(self, row): #puts values of row to list
        rlist = []
        for i in row:
            rlist.append(i.value)
        return rlist

    def set_column_size(self, size): #set width of columns
        for i in self.__columns:
            self.__ws.column_dimensions[str(chr(64 + self.__columns[i]))].width = size

    def rename_column(self, cname, value):
        if cname in self.__columns:
            self.__ws.cell(row=1, column=self.__columns[cname]).value = value

    def del_column(self, cname): #delete column from excel and from columns list
        if cname in self.__columns:
            self.__ws.delete_cols(self.__columns[cname], 1)
            del self.__columns[cname]
            self.__apply_col_del()

    def drop_rows_by_value(self, column, bad_data): #deliting all rows by value in colomn
        if column not in self.__columns:
            return
        i = 2
        dcount = 0
        while True:
            cell_value = self.__ws[str(chr(64 + self.__columns[column]))+str(i)].value
            if cell_value == None:
                break
            if cell_value in bad_data:
                self.del_row(i)
                dcount += 1
            else:
                i += 1
        return dcount

    def keep_rows_by_value(self, column, good_data): #keeping all rows with good_data in colomn
        if column not in self.__columns:
            return
        kcount = 0
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.append(self.__get_list_from_row(self.__ws[1]))
        for i in range(2, self.__ws.max_row + 1):
            cell_value = self.__ws[str(chr(64 + self.__columns[column]))+str(i)].value
            if cell_value in good_data:
                new_ws.append(self.__get_list_from_row(self.__ws[i]))
                kcount += 1
        dcount = self.__ws.max_row - kcount
        self.__wb = new_wb
        self.__ws = new_ws
        self.set_column_size(30)
        return dcount

    def del_row(self, number): #TODO: to this
        self.__ws.delete_rows(number, 1)

    def get_rows_count(self):
        return self.__ws.max_row

    def get_columns_count(self):
        return self.__ws.max_column

    def get_columns(self):
        return self.__columns

    def save(self): #save data to file
        print("------------done------------")
        print(f"saving data to {OUT} ...")
        self.__wb.save(OUT)
        print(f"Total number of rows saved: {self.__ws.max_row}. And total number of columns saved: {self.__ws.max_column}!")
